from fastapi import FastAPI, BackgroundTasks, HTTPException, Form, Request, Depends
from fastapi.responses import HTMLResponse, FileResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
import asyncio
import urllib.parse
import cloudscraper
import random
import logging
from bs4 import BeautifulSoup
from typing import List, Dict, Optional
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import datetime
import os
import uuid
import re
import time
import sqlite3
import jwt
import bcrypt
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from datetime import datetime, timedelta

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

app = FastAPI()
templates = Jinja2Templates(directory="templates")

SECRET_KEY = os.getenv("SECRET_KEY", "your-secret-key-change-in-prod")
ALGORITHM = "HS256"

os.makedirs("downloads", exist_ok=True)

MAX_PAGES = 1000

class SearchRequest(BaseModel):
    url: str
    include_chars: bool = True

class SellerRequest(BaseModel):
    seller_name: str
    include_chars: bool = True

class FavoriteRequest(BaseModel):
    name: str
    urls: List[str]
    include_chars: bool = True

def init_db():
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS users (id INTEGER PRIMARY KEY, username TEXT UNIQUE, password_hash TEXT, status TEXT DEFAULT 'pending')")
    c.execute("CREATE TABLE IF NOT EXISTS favorites (id INTEGER PRIMARY KEY, username TEXT, name TEXT, urls TEXT, created_at TEXT)")
    c.execute("SELECT id FROM users WHERE username=?", ("admin1",))
    if not c.fetchone():
        pw_hash = bcrypt.hashpw("admin33".encode(), bcrypt.gensalt())
        c.execute("INSERT INTO users (username, password_hash, status) VALUES (?, ?, 'admin')", ("admin1", pw_hash))
    else:
        c.execute("UPDATE users SET status='admin' WHERE username='admin1'")
    conn.commit()
    conn.close()

init_db()

def hash_password(password: str) -> bytes:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt())

def verify_password(password: str, hash_bytes: bytes) -> bool:
    return bcrypt.checkpw(password.encode(), hash_bytes)

def create_token(username: str) -> str:
    payload = {"sub": username, "exp": datetime.utcnow() + timedelta(hours=24)}
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(request: Request) -> Optional[Dict[str, str]]:
    token = request.cookies.get("token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if username is None:
            return None
        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        c.execute("SELECT status FROM users WHERE username=?", (username,))
        row = c.fetchone()
        conn.close()
        if row and row[0] in ['accepted', 'admin']:
            return {'username': username, 'status': row[0]}
        return None
    except jwt.PyJWTError:
        return None

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
    'Accept-Language': 'uk-UA,uk;q=0.9,en-US;q=0.8,en;q=0.7',
    'Accept-Encoding': 'gzip, deflate, br',
    'Connection': 'keep-alive',
    'Upgrade-Insecure-Requests': '1'
}

def create_selenium_driver():
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--disable-extensions')
    chrome_options.add_argument('--disable-logging')
    chrome_options.add_argument('--log-level=3')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument(f'user-agent={HEADERS["User-Agent"]}')
    chrome_options.add_experimental_option('excludeSwitches', ['enable-logging', 'enable-automation'])
    chrome_options.add_argument('--disable-blink-features=AutomationControlled')
    chrome_options.add_experimental_option("useAutomationExtension", False)
    

    chrome_options.binary_location = "/nix/store/*/chromium-*/bin/chromium" if os.path.exists("/nix/store") else None
    
    driver = webdriver.Chrome(options=chrome_options)
    driver.set_page_load_timeout(30)
    
    driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
    try:
        driver.execute_cdp_cmd('Network.setUserAgentOverride', {
            "userAgent": HEADERS["User-Agent"]
        })
    except:
        pass
    
    return driver

def wait_for_content_load(driver, timeout=30):
    logging.info("⏳ [Selenium] Очікування загрузки контенту...")
    
    for i in range(timeout):
        time.sleep(1)
        try:
            driver.find_element(By.CSS_SELECTOR, "rz-slider-placeholder")
            if i % 5 == 0:
                logging.info(f"⏳ Placeholder присутствує ({i+1}/{timeout} сек)")
        except NoSuchElementException:
            logging.info(f"✓ Placeholder ісчез після {i+1} сек")
            return True
    
    logging.info("⚠️ Placeholder не ісчез, провіряєм наличіє li елементів...")
    try:
        li_items = driver.find_elements(By.CSS_SELECTOR, "#all_sellers-block li")
        if li_items:
            logging.info(f"✓ Найдены li елементы ({len(li_items)}), продовжаєм")
            return True
    except:
        pass
    
    try:
        offers = driver.find_element(By.CSS_SELECTOR, "rz-product-offers")
        if offers:
            logging.info("✓ Блок rz-product-offers найден, продовжаєм")
            return True
    except:
        pass
    
    logging.warning("⚠️ Контент не загрузився")
    return False

async def fetch_page(session, url, delay=0.5):
    try:
        logging.info(f"Отримання сторінки: {url}")
        response = session.get(url, timeout=15)
        response.raise_for_status()
        await asyncio.sleep(random.uniform(delay, delay + 0.5))
        return response.json().get('data', {})
    except Exception as e:
        logging.error(f"Помилка: {e}")
        return {}

async def fetch_wishlist_count(session, product_id):
    try:
        url = f"https://uss.rozetka.com.ua/session/wishlist/count-goods?country=UA&lang=ua&goods_ids={product_id}"
        response = session.get(url, timeout=10)
        response.raise_for_status()
        await asyncio.sleep(random.uniform(0.1, 0.3))
        json_data = response.json()
        data_array = json_data.get('data', [])
        return data_array[0].get('count', 0) if data_array else 0
    except Exception as e:
        logging.error(f"Помилка wishlist: {e}")
        return 0

async def fetch_product_reviews(session, product_id):
    try:
        url = f"https://rozetka.com.ua/ua/{product_id}/p{product_id}/comments/"
        logging.info(f"Парсинг відгуків товару: {url}")
        
        response = session.get(url, timeout=15)
        response.raise_for_status()
        await asyncio.sleep(random.uniform(0.3, 0.6))
        
        soup = BeautifulSoup(response.text, 'html.parser')
        
        ratings = []
        rating_divs = soup.find_all('div', class_='stars__rating')
        
        for div in rating_divs[:3]:
            style = div.get('style', '')
            match = re.search(r'width:\s*calc\((\d+)%', style)
            if match:
                percent = int(match.group(1))
                stars = percent / 20
                ratings.append(stars)
        
        if len(ratings) >= 3:
            avg = sum(ratings) / len(ratings)
            logging.info(f"✓ Знайдено {len(ratings)} оцінок товару, середня: {avg:.2f}")
            return round(avg, 2)
        elif len(ratings) > 0:
            logging.warning(f"Недостатньо відгуків (знайдено {len(ratings)}, потрібно 3)")
            return None
        else:
            logging.warning(f"Відгуки не знайдено")
            return None
            
    except Exception as e:
        logging.error(f"Помилка парсингу відгуків товару: {e}")
        return None

async def fetch_product_grouping_selenium(product_id, executor):
    try:
        url = f"https://rozetka.com.ua/ua/{product_id}/p{product_id}/"
        logging.info(f"🔍 [Selenium] Початок парсингу групування: {url}")
        
        loop = asyncio.get_event_loop()
        result = await loop.run_in_executor(executor, _selenium_fetch_grouping, url, product_id)
        
        return result
        
    except Exception as e:
        logging.error(f"❌ Помилка парсингу групування: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return {
            'has_grouping': 'Ні',
            'grouping_count': 0,
            'min_price': '',
            'sellers': []
        }

def _selenium_fetch_grouping(url, product_id):
    driver = None
    try:
        driver = create_selenium_driver()
        logging.info(f"🔄 [Selenium] Загрузка страницы...")
        driver.get(url)
        
        time.sleep(3)
        logging.info(f"✓ [Selenium] Страница загружена")
        
        logging.info("📜 [Selenium] Скроллинг к блоку продавцов...")
        try:
            all_sellers_block = driver.find_element(By.CSS_SELECTOR, "#all_sellers-block")
            driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", all_sellers_block)
            time.sleep(2)
            logging.info("✓ [Selenium] Скроллинг выполнен")
        except:
            logging.warning("⚠️ Блок #all_sellers-block не найден для скроллинга")
        
        logging.info("🔘 [Selenium] Поиск кнопки группировки...")
        button_clicked = False
        button_selectors = [
            "rz-toggle-button button",
            "rz-product-offers rz-toggle-button button",
            "#all_sellers-block rz-toggle-button button",
            "button[class*='toggle']"
        ]
        
        for selector in button_selectors:
            try:
                button = driver.find_element(By.CSS_SELECTOR, selector)
                if button.is_displayed():
                    driver.execute_script("arguments[0].scrollIntoView({behavior: 'smooth', block: 'center'});", button)
                    time.sleep(1)
                    
                    driver.execute_script("arguments[0].click();", button)
                    logging.info(f"✓ [Selenium] Кнопка нажата (селектор: {selector})")
                    button_clicked = True
                    
                    time.sleep(3)
                    break
            except NoSuchElementException:
                continue
            except Exception as e:
                logging.warning(f"⚠️ Ошибка при нажатии кнопки ({selector}): {e}")
                continue
        
        if button_clicked:
            logging.info("✓ [Selenium] Кнопка группировки успешно нажата")
        else:
            logging.warning("⚠️ [Selenium] Кнопка группировки не найдена или уже активна")
        
        if not wait_for_content_load(driver, timeout=30):
            logging.warning("⚠️ Контент не загрузился полностью")
        
        time.sleep(2)
        
        selectors = [
            "#all_sellers-block > rz-product-offers > div > ul > li",
            "#all_sellers-block rz-product-offers li",
            "#all_sellers-block li.other-sellers-offers__item",
            "#all_sellers-block li",
        ]
        
        li_items = []
        used_selector = None
        
        for selector in selectors:
            logging.info(f"🔍 [Selenium] Пробуєм селектор: {selector}")
            li_items = driver.find_elements(By.CSS_SELECTOR, selector)
            if li_items:
                used_selector = selector
                logging.info(f"✓ [Selenium] Найдено {len(li_items)} елементів з селектором: {selector}")
                break
            else:
                logging.info(f"  ❌ Селектор не дав результатів")
        
        if not li_items:
            logging.info("ℹ️ [Selenium] Элементы li не найдены - группировка отсутствует")
            
            logging.info("🔍 [Selenium] Отладка - провіряєм структуру страницы:")
            try:
                block = driver.find_element(By.CSS_SELECTOR, "#all_sellers-block")
                html = block.get_attribute('innerHTML')[:1000]
                logging.info(f"📄 HTML блока (первые 1000 символів):\n{html}")
            except:
                logging.error("❌ Блок #all_sellers-block вообще не найден!")
            
            return {
                'has_grouping': 'Ні',
                'grouping_count': 0,
                'min_price': '',
                'sellers': []
            }
        
        prices = []
        sellers = []
        logging.info(f"🔄 [Selenium] Парсинг {len(li_items)} карточек...")
        
        for idx, li in enumerate(li_items, 1):
            try:
                seller_selectors = [
                    "a.other-sellers-offers__seller-link",
                    "a[href*='/seller/']",
                    ".seller-name",
                    "a[class*='seller']"
                ]
                
                seller_name = ''
                for sel in seller_selectors:
                    try:
                        seller_elem = li.find_element(By.CSS_SELECTOR, sel)
                        seller_name = seller_elem.text.strip()
                        if seller_name:
                            sellers.append(seller_name)
                            break
                    except NoSuchElementException:
                        continue
                
                price_selectors = [
                    "p.other-sellers-offers__product-price-main--red",
                    "p.other-sellers-offers__product-price-main",
                    "[class*='price']",
                ]
                
                price_found = False
                for sel in price_selectors:
                    try:
                        price_elem = li.find_element(By.CSS_SELECTOR, sel)
                        price_text = price_elem.text.strip()
                        price_clean = re.sub(r'[^\d]', '', price_text)
                        if price_clean:
                            price_value = float(price_clean)
                            prices.append(price_value)
                            price_found = True
                            break
                    except NoSuchElementException:
                        continue
                        
            except Exception as e:
                logging.error(f"     ❌ Ошибка обработки карточки #{idx}: {e}")
        
        min_price = min(prices) if prices else ''
        logging.info(f"✅ [Selenium] Парсинг завершен:")
        logging.info(f"   - Группировка: Так")
        logging.info(f"   - Количество карточек: {len(li_items)}")
        logging.info(f"   - Минимальная цена: {min_price}")
        logging.info(f"   - Найдено продавцов: {len(sellers)}")
        logging.info(f"   - Найдено цен: {len(prices)}")
        
        return {
            'has_grouping': 'Так',
            'grouping_count': len(li_items),
            'min_price': min_price,
            'sellers': sellers
        }
        
    except Exception as e:
        logging.error(f"❌ [Selenium] Критическая ошибка: {e}")
        import traceback
        logging.error(traceback.format_exc())
        return {
            'has_grouping': 'Ні',
            'grouping_count': 0,
            'min_price': '',
            'sellers': []
        }
    finally:
        if driver:
            driver.quit()

async def fetch_product_page(session, url, executor):
    try:
        loop = asyncio.get_event_loop()
        response = await loop.run_in_executor(executor, lambda: session.get(url, timeout=15))
        response.raise_for_status()
        await asyncio.sleep(random.uniform(0.3, 0.8))
        return response.text
    except Exception as e:
        logging.error(f"Помилка: {e}")
        return None

def parse_characteristics(html: str):
    if not html:
        return {}, ''
    try:
        soup = BeautifulSoup(html, 'html.parser')
        characteristics = {}
        char_lists = soup.find_all('dl', class_='list')
        for char_list in char_lists:
            items = char_list.find_all('div', class_='item')
            for item in items:
                label_elem = item.find('dt', class_='label')
                value_elem = item.find('dd', class_='value')
                if label_elem and value_elem:
                    label = label_elem.get_text(strip=True)
                    values = []
                    sub_list = value_elem.find('ul', class_='sub-list')
                    if sub_list:
                        for li in sub_list.find_all('li'):
                            link = li.find('a')
                            text = link.get_text(strip=True) if link else li.get_text(strip=True)
                            if text:
                                values.append(text)
                    if values:
                        characteristics[label] = ', '.join(values)
        warranty_div = soup.find('div', {'rzhasoverflow': True, 'class': lambda x: x and 'flex-1' in x})
        warranty = warranty_div.get_text(strip=True).replace('\xa0', ' ') if warranty_div else ''
        return characteristics, warranty
    except Exception as e:
        logging.error(f"Помилка парсингу: {e}")
        return {}, ''

async def fetch_delivery_info(session, product_id, price):
    try:
        url = f"https://product-api.rozetka.com.ua/v4/deliveries/get-deliveries?country=UA&lang=ua&city_id=b205dde2-2e2e-4eb9-aef2-a67c82bbdf27&cost={price}&product_id={product_id}"
        response = session.get(url, timeout=15)
        response.raise_for_status()
        await asyncio.sleep(random.uniform(0.2, 0.4))
        data = response.json().get('data', {})
        deliveries = []
        for d in data.get('deliveries', []):
            cost = d.get('cost', {})
            deliveries.append({'title': d.get('title', ''), 'cost': cost.get('new') if cost.get('new') is not None else cost.get('text', 'Н/Д')})
        return {'deliveries': deliveries, 'payments': data.get('payments', '')}
    except Exception as e:
        logging.error(f"Помилка доставки: {e}")
        return {'deliveries': [], 'payments': ''}

async def process_product(session, product, executor, include_chars=True, mode="search"):
    href = product.get('href', '')
    product_id = product.get('id')
    price = product.get('price', 0)
    
    if not href or not product_id:
        return product
    
    wishlist_count = await fetch_wishlist_count(session, product_id)
    characteristics, warranty = {}, ''
    product_avg_rating = None
    grouping_info = None
    
    if include_chars:
        html = await fetch_product_page(session, href, executor)
        characteristics, warranty = parse_characteristics(html)
    
    if mode == "seller" and not include_chars:
        product_avg_rating = await fetch_product_reviews(session, product_id)
        grouping_info = await fetch_product_grouping_selenium(product_id, executor)
    
    delivery_info = await fetch_delivery_info(session, product_id, price) if product_id and price else None
    
    logging.info(f"Оброблено: {product.get('title', '')[:50]}")
    
    result = {
        **product, 
        'characteristics': characteristics, 
        'warranty': warranty,
        'wishlist_count': wishlist_count, 
        'delivery': delivery_info
    }
    
    if mode == "seller" and not include_chars:
        result['product_avg_rating'] = product_avg_rating
        if grouping_info:
            result['has_grouping'] = grouping_info['has_grouping']
            result['grouping_count'] = grouping_info['grouping_count']
            result['min_price_in_group'] = grouping_info['min_price']
            result['sellers_in_group'] = ', '.join(grouping_info.get('sellers', []))
    
    return result

async def fetch_details(session, product_ids):
    try:
        ids_str = ','.join(map(str, product_ids))
        url = f"https://xl-catalog-api.rozetka.com.ua/v4/goods/getDetails?country=UA&lang=ua&goods_group_href=0&product_ids={ids_str}&with_docket=1&with_extra_info=1&with_groups=1"
        detail_headers = {'X-Requested-With': 'XMLHttpRequest'}
        response = session.get(url, headers=detail_headers, timeout=15)
        response.raise_for_status()
        await asyncio.sleep(random.uniform(1, 2))
        return response.json().get('data', [])
    except Exception as e:
        logging.error(f"Помилка деталей: {e}")
        return []

def get_popular_characteristics(products, threshold=350):
    char_count = {}
    for product in products:
        for char_name in product.get('characteristics', {}).keys():
            char_count[char_name] = char_count.get(char_name, 0) + 1
    return [name for name, count in char_count.items() if count >= threshold]

async def export_to_excel(all_products, search_text, filename, include_chars=True, mode="search"):
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    categories = {}
    for product in all_products:
        groups = product.get('groups', [])
        if groups and isinstance(groups, list):
            group_titles = [g.get('title', '') if hasattr(g, 'get') else str(g) for g in groups]
            category = ' / '.join([t for t in group_titles if t])
        else:
            category = ''
        if not category:
            cat = product.get('category', {})
            if hasattr(cat, 'get'):
                category = cat.get('title', 'Без категорії')
            else:
                category = str(cat) if cat else 'Без категорії'
        categories.setdefault(category, []).append(product)
    
    logging.info(f"Знайдено {len(categories)} категорій для розбивки по листам")
    
    for category_name, products in categories.items():
        popular_chars = get_popular_characteristics(products, threshold=350)
        logging.info(f"Створення листа для категорії '{category_name}' ({len(products)} товарів)")
        await create_sheet_with_data(wb, products, search_text, include_chars, popular_chars, 
                                    category_name, mode)
    
    wb.save(filename)
    logging.info(f"Excel файл збережено: {filename}")

async def create_sheet_with_data(wb, products, search_text, include_chars, popular_chars, sheet_base_name, mode):
    unique_chars = set()
    filtered_chars = []
    other_chars = []
    
    if include_chars:
        for product in products:
            unique_chars.update(product.get('characteristics', {}).keys())
        
        popular_chars_set = set(popular_chars) if popular_chars else set()
        filtered_chars = sorted([c for c in unique_chars if c in popular_chars_set])
        other_chars = sorted([c for c in unique_chars if c not in popular_chars_set])
    
    unique_deliveries = set()
    for product in products:
        for d in product.get('delivery', {}).get('deliveries', []):
            if d.get('title'):
                unique_deliveries.add(d['title'])
    unique_deliveries = sorted(list(unique_deliveries))
    
    missing_filtered_chars = False
    if include_chars and filtered_chars:
        for product in products:
            chars = product.get('characteristics', {})
            for filter_char in filtered_chars:
                if not chars.get(filter_char):
                    missing_filtered_chars = True
                    break
            if missing_filtered_chars:
                break
    
    sheet_name = sheet_base_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace(':', '_').replace('[', '_').replace(']', '_')
    
    if include_chars and missing_filtered_chars:
        sheet_name = f"!!!{sheet_name[:28]}"
    
    base_sheet_name = sheet_name
    counter = 1
    while sheet_name in wb.sheetnames:
        sheet_name = f"{base_sheet_name[:28]}_{counter}"
        counter += 1
    
    ws = wb.create_sheet(title=sheet_name)
    
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    dark_green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    fixed_headers = ['Місце в видачі', 'Назва продукта', 'Посилання', 'Пошуковий запит', 'Категорія', 'Бренд', 
                     'Ціна стара', 'Ціна зараз', 'Відгуки зірки', 'Відгуки кількість', 'Кількість в списках бажань', 
                     'Продавець', 'Оплата', 'Гарантія']
    
    if mode == "seller" and not include_chars:
        fixed_headers.extend(['Середня оцінка (перші 3 відгуки)', 'Групування, так/ні', 'Кількість карток у групуванні', 'Мінімальна ціна в групуванні', 'Продавці в групуванні'])
    
    headers = fixed_headers + unique_deliveries
    if include_chars:
        headers += filtered_chars + other_chars
    
    fixed_count = len(fixed_headers)
    delivery_count = len(unique_deliveries)
    filtered_count = len(filtered_chars)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if header in ['Середня оцінка (перші 3 відгуки)', 'Групування, так/ні', 'Кількість карток у групуванні', 'Мінімальна ціна в групуванні', 'Продавці в групуванні']:
            cell.fill = dark_green_fill
        elif col <= fixed_count:
            cell.fill = green_fill
        elif col <= fixed_count + delivery_count:
            cell.fill = orange_fill
        elif col <= fixed_count + delivery_count + filtered_count:
            cell.fill = gray_fill
        else:
            cell.fill = yellow_fill
    
    for idx, product in enumerate(products, 1):
        row = idx + 1
        delivery_dict = {d.get('title', ''): d.get('cost', '') for d in product.get('delivery', {}).get('deliveries', [])}
        
        cat = product.get('category', {})
        if hasattr(cat, 'get'):
            cat_title = cat.get('title', '')
        else:
            cat_title = str(cat) if cat else ''
        
        data = [
            idx, product.get('title', ''), product.get('href', ''), search_text,
            cat_title, product.get('brand', ''),
            product.get('old_price', ''), product.get('price', ''),
            product.get('comments_mark', ''), product.get('comments_amount', 0),
            product.get('wishlist_count', 0), product.get('seller', {}).get('title', ''),
            product.get('delivery', {}).get('payments', ''), product.get('warranty', '')
        ]
        
        if mode == "seller" and not include_chars:
            data.append(product.get('product_avg_rating', ''))
            data.append(product.get('has_grouping', ''))
            data.append(product.get('grouping_count', ''))
            data.append(product.get('min_price_in_group', ''))
            data.append(product.get('sellers_in_group', ''))
        
        for delivery_name in unique_deliveries:
            data.append(delivery_dict.get(delivery_name, ''))
        
        if include_chars:
            chars = product.get('characteristics', {})
            for char_key in filtered_chars + other_chars:
                data.append(chars.get(char_key, ''))
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)

def extract_product_ids_from_urls(urls: List[str]) -> List[int]:
    """Витягує ID товарів з URL"""
    product_ids = []
    for url in urls:
        match = re.search(r'/p(\d+)/', url)
        if match:
            product_ids.append(int(match.group(1)))
    return product_ids

@app.get("/", response_class=HTMLResponse)
async def root(request: Request, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    css = """
    <style>
    body { font-family: Arial; max-width: 800px; margin: 50px auto; padding: 20px; background: #ffffff; color: #333; }
    h1 { text-align: center; color: #333; }
    .option { background: #f5f5f5; padding: 20px; margin: 20px 0; border-radius: 8px; }
    input, textarea { width: 100%; padding: 10px; margin: 10px 0; box-sizing: border-box; }
    textarea { min-height: 100px; font-family: monospace; }
    button { background: #32CD32; color: white; border: none; padding: 12px 20px; 
             border-radius: 4px; cursor: pointer; width: 100%; font-size: 16px; margin: 5px 0; }
    button:hover { background: #228B22; }
    button.secondary { background: #4169E1; }
    button.secondary:hover { background: #1E90FF; }
    button.danger { background: #DC143C; }
    button.danger:hover { background: #B22222; }
    .checkbox { width: auto; margin-right: 10px; }
    label { display: flex; align-items: center; margin: 10px 0; }
    #status { margin-top: 20px; padding: 10px; background: #e3f2fd; border-radius: 4px; display: none; }
    .auth-form { background: #f5f5f5; padding: 20px; margin: 20px 0; border-radius: 8px; }
    .favorites-list { max-height: 300px; overflow-y: auto; margin: 10px 0; }
    .favorite-item { background: white; padding: 10px; margin: 5px 0; border-radius: 4px; display: flex; justify-content: space-between; align-items: center; }
    .favorite-item button { width: auto; margin: 0 5px; padding: 5px 15px; }
    select { width: 100%; padding: 10px; margin: 10px 0; }
    </style>
    """
    if current_user:
        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        c.execute("SELECT id, name, urls, created_at FROM favorites WHERE username=?", (current_user['username'],))
        favorites = c.fetchall()
        conn.close()
        
        favorites_html = ""
        if favorites:
            favorites_html = '<div class="favorites-list">'
            for fav_id, name, urls_json, created_at in favorites:
                urls = json.loads(urls_json)
                favorites_html += f'''
                <div class="favorite-item">
                    <div>
                        <strong>{name}</strong><br>
                        <small>{len(urls)} товарів | {created_at}</small>
                    </div>
                    <div>
                        <button class="secondary" onclick="runFavorite({fav_id})">Парсити</button>
                        <button class="danger" onclick="deleteFavorite({fav_id})">Видалити</button>
                    </div>
                </div>
                '''
            favorites_html += '</div>'
        
        html = f"""
        {css}
        <!DOCTYPE html>
        <html>
        <head><title>Rozetka Parser</title><meta charset="utf-8"></head>
        <body>
            <h1>Rozetka Parser (швидкий режим - перші 2 сторінки)</h1>
            <p>Добро пожаловать, {current_user['username']}!</p>
            
            <div class="option">
                <h2>1. Парсинг по запиту/категорії</h2>
                <input type="text" id="searchUrl" placeholder="URL пошуку">
                <label><input type="checkbox" class="checkbox" id="searchChars" checked> З характеристиками</label>
                <button onclick="runSearch()">Запустити</button>
            </div>
            
            <div class="option">
                <h2>2. Парсинг продавця</h2>
                <input type="text" id="sellerName" placeholder="Назва або URL продавця">
                <label><input type="checkbox" class="checkbox" id="sellerChars" checked> З характеристиками</label>
                <button onclick="runSeller()">Запустити</button>
            </div>
            
            <div class="option">
                <h2>3. Обрані товари</h2>
                <input type="text" id="favoriteName" placeholder="Назва списку">
                <textarea id="favoriteUrls" placeholder="Посилання на товари (кожне з нового рядка)&#10;Приклад:&#10;https://rozetka.com.ua/ua/product/p123456/&#10;https://rozetka.com.ua/ua/product/p789012/"></textarea>
                <label><input type="checkbox" class="checkbox" id="favoriteChars" checked> З характеристиками</label>
                <button onclick="saveFavorite()">Зберегти список</button>
                <button class="secondary" onclick="runFavoriteQuick()">Парсити без збереження</button>
                
                <h3>Збережені списки:</h3>
                {favorites_html if favorites_html else '<p>Немає збережених списків</p>'}
            </div>
            
            <div id="status"></div>
            
            <button onclick="window.location.href='/admin'">Адмін панель</button>
            <button onclick="logout()">Вийти</button>
            
            <script>
                function showStatus(msg) {{
                    const status = document.getElementById('status');
                    status.textContent = msg;
                    status.style.display = 'block';
                }}
                
                async function runSearch() {{
                    const url = document.getElementById('searchUrl').value;
                    const includeChars = document.getElementById('searchChars').checked;
                    if (!url) {{ alert('Введіть URL'); return; }}
                    
                    showStatus('Обробка...');
                    const res = await fetch('/api/search', {{
                        method: 'POST',
                        headers: {{'Content-Type': 'application/json'}},
                        body: JSON.stringify({{url, include_chars: includeChars}})
                    }});
                    const data = await res.json();
                    if (data.filename) {{
                        showStatus('Готово!');
                        window.location.href = '/download/' + data.filename;
                    }} else {{
                        showStatus('Помилка: ' + data.error);
                    }}
                }}
                
                async function runSeller() {{
                    let sellerName = document.getElementById('sellerName').value;
                    const includeChars = document.getElementById('sellerChars').checked;
                    if (!sellerName) {{ alert('Введіть назву продавця'); return; }}
                    
                    if (sellerName.includes('rozetka.com.ua')) {{
                        const parts = sellerName.split('/seller/');
                        if (parts.length > 1) {{
                            sellerName = parts[1].split('/')[0];
                        }}
                    }}
                    
                    showStatus('Обробка...');
                    const res = await fetch('/api/seller', {{
                        method: 'POST',
                        headers: {{'Content-Type': 'application/json'}},
                        body: JSON.stringify({{seller_name: sellerName, include_chars: includeChars}})
                    }});
                    const data = await res.json();
                    if (data.filename) {{
                        showStatus('Готово!');
                        window.location.href = '/download/' + data.filename;
                    }} else {{
                        showStatus('Помилка: ' + data.error);
                    }}
                }}
                
                async function saveFavorite() {{
                    const name = document.getElementById('favoriteName').value;
                    const urlsText = document.getElementById('favoriteUrls').value;
                    const includeChars = document.getElementById('favoriteChars').checked;
                    
                    if (!name || !urlsText) {{
                        alert('Введіть назву та посилання');
                        return;
                    }}
                    
                    const urls = urlsText.split('\\n').filter(u => u.trim());
                    if (urls.length === 0) {{
                        alert('Немає валідних посилань');
                        return;
                    }}
                    
                    showStatus('Збереження...');
                    const res = await fetch('/api/favorites/save', {{
                        method: 'POST',
                        headers: {{'Content-Type': 'application/json'}},
                        body: JSON.stringify({{name, urls, include_chars: includeChars}})
                    }});
                    const data = await res.json();
                    if (data.success) {{
                        showStatus('Збережено!');
                        setTimeout(() => location.reload(), 1000);
                    }} else {{
                        showStatus('Помилка: ' + data.error);
                    }}
                }}
                
                async function runFavoriteQuick() {{
                    const urlsText = document.getElementById('favoriteUrls').value;
                    const includeChars = document.getElementById('favoriteChars').checked;
                    
                    if (!urlsText) {{
                        alert('Введіть посилання');
                        return;
                    }}
                    
                    const urls = urlsText.split('\\n').filter(u => u.trim());
                    if (urls.length === 0) {{
                        alert('Немає валідних посилань');
                        return;
                    }}
                    
                    showStatus('Обробка...');
                    const res = await fetch('/api/favorites/parse', {{
                        method: 'POST',
                        headers: {{'Content-Type': 'application/json'}},
                        body: JSON.stringify({{urls, include_chars: includeChars}})
                    }});
                    const data = await res.json();
                    if (data.filename) {{
                        showStatus('Готово!');
                        window.location.href = '/download/' + data.filename;
                    }} else {{
                        showStatus('Помилка: ' + data.error);
                    }}
                }}
                
                async function runFavorite(favoriteId) {{
                    showStatus('Обробка...');
                    const res = await fetch('/api/favorites/parse/' + favoriteId, {{
                        method: 'POST'
                    }});
                    const data = await res.json();
                    if (data.filename) {{
                        showStatus('Готово!');
                        window.location.href = '/download/' + data.filename;
                    }} else {{
                        showStatus('Помилка: ' + data.error);
                    }}
                }}
                
                async function deleteFavorite(favoriteId) {{
                    if (!confirm('Видалити цей список?')) return;
                    
                    const res = await fetch('/api/favorites/delete/' + favoriteId, {{
                        method: 'DELETE'
                    }});
                    const data = await res.json();
                    if (data.success) {{
                        location.reload();
                    }} else {{
                        alert('Помилка: ' + data.error);
                    }}
                }}
                
                function logout() {{
                    document.cookie = 'token=; expires=Thu, 01 Jan 1970 00:00:00 UTC; path=/;';
                    window.location.href = '/';
                }}
            </script>
        </body>
        </html>
        """
        return HTMLResponse(content=html)
    
    html = f"""
    {css}
    <!DOCTYPE html>
    <html>
    <head><title>Rozetka Parser Auth</title><meta charset="utf-8"></head>
    <body>
        <h1>Rozetka Parser</h1>
        <div class="auth-form">
            <h2>Реєстрація</h2>
            <form method="post" action="/register">
                <input name="username" placeholder="Логін"><br>
                <input name="password" type="password" placeholder="Пароль"><br>
                <button>Зареєструватись</button>
            </form>
        </div>
        <div class="auth-form">
            <h2>Вхід</h2>
            <form method="post" action="/login">
                <input name="username" placeholder="Логін"><br>
                <input name="password" type="password" placeholder="Пароль"><br>
                <button>Войти</button>
            </form>
        </div>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@app.post("/register")
async def register(username: str = Form(), password: str = Form()):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT id FROM users WHERE username=?", (username,))
    if c.fetchone():
        conn.close()
        raise HTTPException(400, "Користувач існує")
    pw_hash = hash_password(password)
    c.execute("INSERT INTO users (username, password_hash, status) VALUES (?, ?, 'pending')", (username, pw_hash))
    conn.commit()
    conn.close()
    css = """
    <style>
    body { background: #ffffff; color: #333; font-family: Arial; max-width: 600px; margin: 50px auto; padding: 20px; text-align: center; }
    button { background: #32CD32; color: white; border: none; padding: 12px 20px; margin: 10px; cursor: pointer; border-radius: 4px; font-size: 16px; }
    button:hover { background: #228B22; }
    .status { background: #f0f0f0; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #32CD32; }
    .loader { border: 5px solid #f3f3f3; border-top: 5px solid #32CD32; border-radius: 50%; width: 50px; height: 50px; animation: spin 1s linear infinite; margin: 20px auto; }
    @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
    </style>
    """
    html = f"""
    {css}
    <!DOCTYPE html>
    <html>
    <head><title>Реєстрація</title><meta charset="utf-8"></head>
    <body>
        <h1>✅ Запит відправлено на модерацію</h1>
        <div class="status">
            <p>Ваш обліковий запис очікує схвалення адміністратора</p>
            <p>Сторінка автоматично оновиться після схвалення</p>
        </div>
        <div class="loader"></div>
        <p id="timer">Перевірка через 5 секунд...</p>
        <button onclick="checkNow()">Перевірити зараз</button>
        <button onclick="window.location.href='/'">На головну</button>
        
        <script>
            let countdown = 5;
            let checkInterval;
            const username = '{username}';
            const password = '{password}';
            
            function updateTimer() {{
                document.getElementById('timer').textContent = `Перевірка через ${{countdown}} секунд...`;
                countdown--;
                if (countdown < 0) {{
                    countdown = 5;
                    checkStatus();
                }}
            }}
            
            async function checkStatus() {{
                try {{
                    const response = await fetch('/check-status/' + username);
                    const data = await response.json();
                    
                    if (data.status === 'accepted') {{
                        document.getElementById('timer').textContent = '✅ Схвалено! Виконується вхід...';
                        
                        const loginResponse = await fetch('/auto-login', {{
                            method: 'POST',
                            headers: {{'Content-Type': 'application/json'}},
                            body: JSON.stringify({{username: username, password: password}})
                        }});
                        
                        if (loginResponse.ok) {{
                            const loginData = await loginResponse.json();
                            document.cookie = 'token=' + loginData.token + '; path=/; max-age=86400';
                            
                            setTimeout(() => {{
                                window.location.href = '/';
                            }}, 500);
                        }} else {{
                            document.getElementById('timer').textContent = '❌ Помилка входу';
                        }}
                    }} else if (data.status === 'rejected') {{
                        document.getElementById('timer').textContent = '❌ Реєстрацію відхилено';
                        clearInterval(checkInterval);
                    }}
                }} catch (e) {{
                    console.error('Помилка перевірки статусу:', e);
                }}
            }}
            
            function checkNow() {{
                countdown = 0;
                checkStatus();
                countdown = 5;
            }}
            
            checkInterval = setInterval(updateTimer, 1000);
            setTimeout(checkStatus, 5000);
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html)

@app.get("/check-status/{username}")
async def check_status(username: str):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT status FROM users WHERE username=?", (username,))
    row = c.fetchone()
    conn.close()
    if row:
        return {"status": row[0]}
    return {"status": "not_found"}

@app.post("/auto-login")
async def auto_login(request: Request):
    data = await request.json()
    username = data.get('username')
    password = data.get('password')
    
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT password_hash, status FROM users WHERE username=?", (username,))
    row = c.fetchone()
    conn.close()
    
    if not row or not verify_password(password, row[0]):
        raise HTTPException(400, "Невірні дані")
    
    status = row[1]
    if status not in ['accepted', 'admin']:
        raise HTTPException(400, "Очікується схвалення")
    
    token = create_token(username)
    return {"token": token}

@app.post("/login")
async def login(username: str = Form(), password: str = Form()):
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT password_hash, status FROM users WHERE username=?", (username,))
    row = c.fetchone()
    conn.close()
    if not row or not verify_password(password, row[0]):
        raise HTTPException(400, "Невірні дані")
    status = row[1]
    if status not in ['accepted', 'admin']:
        if status == 'pending':
            raise HTTPException(400, "Реєстрація очікує схвалення адміна")
        raise HTTPException(400, "Реєстрація відхилена")
    token = create_token(username)
    response = RedirectResponse(url="/", status_code=303)
    response.set_cookie(key="token", value=token, httponly=True)
    return response

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user or current_user['username'] != "admin1":
        raise HTTPException(403, "Тільки для адміна")
    css = "<style>body { background: #ffffff; color: #333; } button { background: #32CD32; color: white; border: none; padding: 10px; margin: 5px; cursor: pointer; border-radius: 4px; } button:hover { background: #228B22; }</style>"
    html = f"""
    {css}
    <!DOCTYPE html>
    <html>
    <head><title>Адмін</title></head>
    <body>
        <h1>Адмін панель</h1>
        <p>Користувачі:</p>
        <ul>
    """
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("SELECT username, status FROM users WHERE status != 'admin'")
    for row in c.fetchall():
        username, status = row
        html += f"<li>{username} ({status}) "
        if status == 'pending':
            html += f'<form method="post" action="/accept/{username}" style="display:inline;"><button>Прийняти</button></form> '
            html += f'<form method="post" action="/reject/{username}" style="display:inline;"><button>Відхилити</button></form> '
        html += f'<form method="post" action="/delete/{username}" style="display:inline;"><button>Видалити</button></form></li>'
    html += """
        </ul>
        <button onclick="window.location.href='/'">Головна</button>
    </body>
    </html>
    """
    conn.close()
    return HTMLResponse(content=html)

@app.post("/accept/{username}")
async def accept_user(username: str, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user or current_user['username'] != "admin1":
        raise HTTPException(403, "Тільки для адміна")
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("UPDATE users SET status='accepted' WHERE username=?", (username,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)

@app.post("/reject/{username}")
async def reject_user(username: str, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user or current_user['username'] != "admin1":
        raise HTTPException(403, "Тільки для адміна")
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("UPDATE users SET status='rejected' WHERE username=?", (username,))
    conn.commit()
    conn.close()
    return RedirectResponse(url="/admin", status_code=303)

@app.post("/delete/{username}")
async def delete_user(username: str, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user or current_user['username'] != "admin1":
        raise HTTPException(403, "Тільки для адміна")
    response = RedirectResponse(url="/admin", status_code=303)
    if username == current_user['username']:
        response = RedirectResponse(url="/", status_code=303)
        response.delete_cookie("token")
    conn = sqlite3.connect("users.db")
    c = conn.cursor()
    c.execute("DELETE FROM users WHERE username=?", (username,))
    conn.commit()
    conn.close()
    return response

@app.post("/api/favorites/save")
async def save_favorite(req: FavoriteRequest, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(401, "Не авторизовано")
    try:
        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        urls_json = json.dumps(req.urls)
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
        c.execute("INSERT INTO favorites (username, name, urls, created_at) VALUES (?, ?, ?, ?)",
                  (current_user['username'], req.name, urls_json, created_at))
        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        logging.error(f"Помилка збереження: {e}")
        return {"success": False, "error": str(e)}

@app.post("/api/favorites/parse")
async def parse_favorite_quick(request: Request, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(401, "Не авторизовано")
    try:
        data = await request.json()
        urls = data.get('urls', [])
        include_chars = data.get('include_chars', True)
        
        product_ids = extract_product_ids_from_urls(urls)
        if not product_ids:
            raise HTTPException(400, "Не знайдено валідних ID товарів")
        
        session = cloudscraper.create_scraper()
        session.headers.update(HEADERS)
        
        executor = ThreadPoolExecutor(max_workers=20)
        all_products = []
        
        batch_size = 60
        for i in range(0, len(product_ids), batch_size):
            batch = product_ids[i:i + batch_size]
            details = await fetch_details(session, batch)
            tasks = [process_product(session, p, executor, include_chars, "favorites") for p in details]
            batch_results = await asyncio.gather(*tasks)
            all_products.extend(batch_results)
        
        executor.shutdown(wait=True)
        
        filename = f"downloads/rozetka_favorites_{uuid.uuid4().hex[:8]}.xlsx"
        await export_to_excel(all_products, "Обрані товари", filename, include_chars, "favorites")
        
        return {"filename": os.path.basename(filename), "count": len(all_products)}
    except Exception as e:
        logging.error(f"Помилка: {e}")
        raise HTTPException(500, str(e))

@app.post("/api/favorites/parse/{favorite_id}")
async def parse_favorite(favorite_id: int, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(401, "Не авторизовано")
    try:
        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        c.execute("SELECT name, urls FROM favorites WHERE id=? AND username=?", (favorite_id, current_user['username']))
        row = c.fetchone()
        conn.close()
        
        if not row:
            raise HTTPException(404, "Список не знайдено")
        
        name, urls_json = row
        urls = json.loads(urls_json)
        
        product_ids = extract_product_ids_from_urls(urls)
        if not product_ids:
            raise HTTPException(400, "Не знайдено валідних ID товарів")
        
        session = cloudscraper.create_scraper()
        session.headers.update(HEADERS)
        
        executor = ThreadPoolExecutor(max_workers=20)
        all_products = []
        
        batch_size = 60
        for i in range(0, len(product_ids), batch_size):
            batch = product_ids[i:i + batch_size]
            details = await fetch_details(session, batch)
            tasks = [process_product(session, p, executor, True, "favorites") for p in details]
            batch_results = await asyncio.gather(*tasks)
            all_products.extend(batch_results)
        
        executor.shutdown(wait=True)
        
        filename = f"downloads/rozetka_{name.replace(' ', '_')}_{uuid.uuid4().hex[:8]}.xlsx"
        await export_to_excel(all_products, name, filename, True, "favorites")
        
        return {"filename": os.path.basename(filename), "count": len(all_products)}
    except Exception as e:
        logging.error(f"Помилка: {e}")
        raise HTTPException(500, str(e))

@app.delete("/api/favorites/delete/{favorite_id}")
async def delete_favorite(favorite_id: int, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(401, "Не авторизовано")
    try:
        conn = sqlite3.connect("users.db")
        c = conn.cursor()
        c.execute("DELETE FROM favorites WHERE id=? AND username=?", (favorite_id, current_user['username']))
        conn.commit()
        conn.close()
        return {"success": True}
    except Exception as e:
        logging.error(f"Помилка: {e}")
        return {"success": False, "error": str(e)}

@app.post("/api/search")
async def api_search(req: SearchRequest, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(401, "Не авторизовано")
    try:
        query = urllib.parse.urlparse(req.url).query
        text = urllib.parse.parse_qs(query).get('text', [''])[0]
        if not text:
            raise HTTPException(400, "Не знайдено параметр 'text'")
        
        base_url = "https://search.rozetka.com.ua/ua/search/api/v7/?country=UA&lang=ua&text=" + urllib.parse.quote(text)
        
        session = cloudscraper.create_scraper()
        session.headers.update(HEADERS)
        
        data = await fetch_page(session, base_url)
        total_pages = min(data.get('pagination', {}).get('total_pages', 1), MAX_PAGES)
        total_found = data.get('pagination', {}).get('total_found', 0)
        logging.info(f"Знайдено товарів: {total_found}, Парсимо перші {total_pages} сторінок")
        
        all_product_ids = []
        for page in range(1, total_pages + 1):
            page_url = f"{base_url}&page={page}"
            data = await fetch_page(session, page_url)
            page_ids = [p.get('id') for p in data.get('goods', []) if p.get('id')]
            if not page_ids:
                logging.warning(f"Сторінка {page} порожня, зупиняємо парсинг")
                break
            all_product_ids.extend(page_ids)
            logging.info(f"Сторінка {page}/{total_pages}: зібрано {len(page_ids)} товарів (всього: {len(all_product_ids)})")
        
        logging.info(f"Всього товарів: {len(all_product_ids)}")
        
        executor = ThreadPoolExecutor(max_workers=50)
        all_products = []
        
        batch_size = 60
        for i in range(0, len(all_product_ids), batch_size):
            batch = all_product_ids[i:i + batch_size]
            details = await fetch_details(session, batch)
            tasks = [process_product(session, p, executor, req.include_chars, "search") for p in details]
            batch_results = await asyncio.gather(*tasks)
            all_products.extend(batch_results)
        
        executor.shutdown(wait=True)
        
        filename = f"downloads/rozetka_search_{text[:20].replace(' ', '_')}_{uuid.uuid4().hex[:8]}.xlsx"
        await export_to_excel(all_products, text, filename, req.include_chars, "search")
        
        return {"filename": os.path.basename(filename), "count": len(all_products)}
    except Exception as e:
        logging.error(f"Помилка: {e}")
        raise HTTPException(500, str(e))

@app.post("/api/seller")
async def api_seller(req: SellerRequest, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(401, "Не авторизовано")
    try:
        async def fetch_seller_api(session, seller_name, page=1):
            url = f"https://search.rozetka.com.ua/ua/seller/api/v7/?front-type=xl&country=UA&lang=ua&name={seller_name}&page={page}"
            response = session.get(url, timeout=15)
            response.raise_for_status()
            await asyncio.sleep(random.uniform(0.2, 0.4))
            data = response.json().get('data', {})
            return {
                'seller_title': data.get('seller_info', {}).get('title', ''),
                'product_ids': [item.get('id') for item in data.get('goods', []) if item.get('id')],
                'total_pages': data.get('pagination', {}).get('total_pages', 1)
            }
        
        session = cloudscraper.create_scraper()
        session.headers.update(HEADERS)
        
        first_page = await fetch_seller_api(session, req.seller_name, 1)
        seller_title = first_page['seller_title']
        total_pages = min(first_page['total_pages'], MAX_PAGES)
        all_product_ids = first_page['product_ids']
        
        logging.info(f"Продавець: {seller_title}, Парсимо перші {total_pages} сторінок, Перша сторінка: {len(all_product_ids)} товарів")
        
        for page in range(2, total_pages + 1):
            page_data = await fetch_seller_api(session, req.seller_name, page)
            if not page_data['product_ids']:
                logging.warning(f"Сторінка {page} порожня, зупиняємо парсинг")
                break
            all_product_ids.extend(page_data['product_ids'])
            logging.info(f"Сторінка {page}/{total_pages}: зібрано {len(page_data['product_ids'])} товарів (всього: {len(all_product_ids)})")
        
        logging.info(f"Всього товарів: {len(all_product_ids)}")
        
        executor = ThreadPoolExecutor(max_workers=20)
        all_products = []
        
        batch_size = 60
        for i in range(0, len(all_product_ids), batch_size):
            batch = all_product_ids[i:i + batch_size]
            details = await fetch_details(session, batch)
            tasks = [process_product(session, p, executor, req.include_chars, "seller") for p in details]
            batch_results = await asyncio.gather(*tasks)
            all_products.extend(batch_results)
        
        executor.shutdown(wait=True)
        
        filename = f"downloads/rozetka_seller_{req.seller_name[:20].replace(' ', '_')}_{uuid.uuid4().hex[:8]}.xlsx"
        await export_to_excel(all_products, seller_title, filename, req.include_chars, "seller")
        
        return {"filename": os.path.basename(filename), "count": len(all_products)}
    except Exception as e:
        logging.error(f"Помилка: {e}")
        raise HTTPException(500, str(e))

@app.get("/download/{filename}")
async def download_file(filename: str, current_user: Optional[Dict[str, str]] = Depends(get_current_user)):
    if not current_user:
        raise HTTPException(401, "Не авторизовано")
    file_path = f"downloads/{filename}"
    if os.path.exists(file_path):
        return FileResponse(file_path, filename=filename, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    raise HTTPException(404, "Файл не знайдено")

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)

